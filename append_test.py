
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font, Alignment

full_wm = ['2023-03-22', '월드와이드메모리', '1,000', '4,000', '25,000', '2,000', '12,000', '40,000', '13,000', '30,000', '67,000', '17,000', '45,000', '100,000', '50,000', '66,000', '130,000', '65,000', '95,000', '200,000', '1,000', '11,000', '4,000', '11,000', '26,000', '5,000', '15,000', '20,000', '25,000', '40,000', None]
full_nm = ['2023-03-22', '나노메모리', 0, '4,000', '25,000', 0, '12,000', '40,000', '13,000', '30,000', '70,000', '20,000', '45,000', '10,000', '50,000', '68,000', '132,000', '65,000', '90,000', '175,000', '1,000', '11,000', '26,000', '11,000', '4,000', 0, '15,000', '25,000', '30,000', 0, None]
full_danawa = ['2023-03-22', '다나와', '11,630', '31,400', '36,630', '75,000', '84,850', '122,100', '15,590', '27,990', '52,500', '79,200', '123,470', '131,150', '57,940', '81,200', '116,990', '177,510', '217,000', '297,200', '4,400', '21,080', '13,000', '22,620', '48,840', 0, 0, 0, 0, 0, None]

# 함수 지정

# 3개의 리스트에서 '30,000' → 30000 (str → int )
def change_int(list):
    # 리스트의 각 요소에서 ',' 제거
    for i in range(len(list)):
        if list[i] is not None and not isinstance(list[i], int):  # None이 아닌 경우에만
            list[i] = list[i].replace(',', '')  # ',' 제거

    # 숫자 str → 숫자 int
    for j in range(2, len(list)-1):
        list[j] = int(list[j])  # int형으로 변환


change_int(full_wm)  # 월드메모리 리스트
change_int(full_nm)  # 나노메모리 리스트
change_int(full_danawa)  # 다나와 리스트

d = ['2023-03-20', full_danawa[5], 'i3 변동' , full_danawa[11], 'i5 변동', full_danawa[17], None]

e = ['2023-03-20', full_danawa[20], '4g ram 변동', full_danawa[23], '8g ram 변동', None]


# 엑셀 파일 열기
wb = openpyxl.load_workbook('it_data2.xlsx')

# 1번 시트 선택
ws = wb['1. IT 전체 개별단가']

# C열의 셀들 가져오기
cells = ws['C']

# # C열의 마지막 행 찾기
# last_row = max(c.row for c in cells if c.value is not None)

# 마지막 행 번호 구하기
last_row = ws.max_row

# 마지막 행 + 2 아래 행 삭제
ws.delete_rows(last_row-1)

# last_row = ws.max_row

# 리스트의 요소를 하나씩 가져와서 B열의 마지막 행 다음 셀에 넣기
for i in range(len(full_wm)):
    ws.cell(last_row -2, i + 3).value = full_wm[i]

for i in range(len(full_nm)):
    ws.cell(last_row - 1, i + 3).value = full_nm[i]

for i in range(len(full_danawa)):
    ws.cell(last_row, i + 3).value = full_danawa[i]
    
    
# 표 아래에 참고 내용 입력
ws.cell(last_row + 3, 33).value = '*빨간음영 : 주력모델'  # 글자 데이터 입력
ws.cell(last_row + 3, 33).font = Font(color='FF0000', bold=True)  # 색상, 볼드체
ws.cell(last_row + 3, 33).alignment = align = Alignment(horizontal='right')  # 우측 정렬 

# # 2번 시트
ws = wb['2. IT 주력모델 시세현황']


# B열의 셀들 가져오기
cells = ws['B']


# 마지막행 B열 값 확인
last_row = ws.max_row

# B열 (넘버링) 위의 행 번호 + 1 입력 
last_value = ws.cell(row=last_row, column=2).value  # 마지막 값 불러오기
ws.cell(row=last_row+1, column=2, value=last_value + 1)

# M열 (넘버링) 위의 행 번호 + 1 입력
last_value = ws.cell(row=last_row, column=13).value
ws.cell(row=last_row+1, column=13, value=last_value + 1)


for i in range(len(d)):
    ws.cell(last_row + 1, i+3).value = d[i]
    
for i in range(len(e)):
    ws.cell(last_row + 1, i+14).value = e[i]
    
# ws.cell(row=last_row + 1, column=5, value='-') 



for k in [4,6,8,15,17]:
    # k열의 마지막행과 마지막행의 윗행 값 획득
    last_row_val = ws.cell(row=last_row+1, column=k).value
    prev_row_val = ws.cell(row=last_row, column=k).value

    # 마지막 행의 k+1 열에 계산값(변동액) 입력
    ws.cell(row=last_row+1, column= k+1).value = last_row_val - prev_row_val


# # 변경사항 저장
# wb.save('example2.xlsx')

# 변경된 내용을 저장
wb.save('example.xlsx')
