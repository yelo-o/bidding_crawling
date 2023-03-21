from openpyxl import load_workbook

# 워크북 불러오기
wb = load_workbook('example.xlsx')

# 워크시트 선택
ws = wb.active

# A1 셀의 서식을 B1 셀에 복사
ws['B1'].font = ws['A1'].font.copy()
ws['B1'].alignment = ws['A1'].alignment.copy()
ws['B1'].border = ws['A1'].border.copy()
ws['B1'].fill = ws['A1'].fill.copy()
ws['B1'].number_format = ws['A1'].number_format.copy()

# 변경사항 저장
wb.save('example.xlsx')