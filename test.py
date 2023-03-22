import openpyxl

# 엑셀 파일 열기
wb = openpyxl.load_workbook('파일명.xlsx')

# 현재 활성화된 시트 선택
ws = wb.active

# 마지막 행의 3번째 열 값과 바로 윗 행의 3번째 열 값 구하기
last_row = ws.max_row
last_col = ws.max_column
last_row_val = ws.cell(row=last_row, column=3).value
prev_row_val = ws.cell(row=last_row-1, column=3).value

# 마지막 행의 4번째 열에 값 입력하기
ws.cell(row=last_row, column=4).value = last_row_val - prev_row_val

# 변경된 내용 저장하기
wb.save('파일명.xlsx')
