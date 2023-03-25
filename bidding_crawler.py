# from openpyxl import Workbook
import os
from gui import file_path

# 파일선택창에서 아무것도 선택하지 않고 창을 닫을 시에 프로그램 종료
if not file_path:
    exit()

from datetime import datetime

from danawa import danawa_prices  # 다나와
from nanoMemory import new_li_prices  # 나노메모리
from worldMemory import world_total_prices  # 월드메모리
import openpyxl
from openpyxl.styles import Font, Alignment

# 함수 지정

# 3개의 리스트에서 '30,000' → 30000 (str → int )
def change_int(list):
    # 리스트의 각 요소에서 ',' 제거
    for i in range(len(list)):
        if list[i] is not None and not isinstance(list[i], int):  # None이 아닌 경우에만
            list[i] = list[i].replace(',', '')  # ',' 제거

    # 숫자 str → 숫자 int
    for j in range(2, len(list)-1):
        list[j] = int(list[j])  # int형으로 변환
        
now = datetime.now()
cr_today = now.strftime("%Y-%m-%d")

full_wm = [cr_today, '월드와이드메모리'] + world_total_prices + [None]
full_nm = [cr_today, '나노메모리'] + new_li_prices + [None]
full_danawa = [cr_today, '다나와'] + danawa_prices + [None]

change_int(full_wm)  # 월드메모리 리스트
change_int(full_nm)  # 나노메모리 리스트
change_int(full_danawa)  # 다나와 리스트


d = [cr_today, full_danawa[5], 'i3 변동' , full_danawa[11], 'i5 변동', full_danawa[17], None]
e = [cr_today, full_danawa[20], '4g ram 변동', full_danawa[23], '8g ram 변동', None]

print(full_wm, len(full_wm))
print(full_nm, len(full_nm))
print(full_danawa, len(full_danawa))


wb = openpyxl.load_workbook('it_data.xlsx')  # 원본
wb = openpyxl.load_workbook(filename=file_path)  # 테스트

# 1번 시트 선택
ws = wb['1. IT 전체 개별단가']

# # C열의 셀들 가져오기
cells = ws['C']


# # C열의 마지막 행 찾기
# last_row = max(c.row for c in cells if c.value is not None)

# # 마지막 행 번호 구하기
last_row = ws.max_row

# # 마지막 행 + 2 아래 행 삭제
ws.delete_rows(last_row-1)

# last_row = ws.max_row

# 리스트의 요소를 하나씩 가져와서 B열의 마지막 행 다음 셀에 넣기
for i in range(len(full_wm)):
    ws.cell(last_row -2, i + 3).value = full_wm[i]

for i in range(len(full_nm)):
    ws.cell(last_row - 1, i + 3).value = full_nm[i]

for i in range(len(full_danawa)):
    ws.cell(last_row, i + 3).value = full_danawa[i]
    
    
# 표 아래에 참고 내용 입력
ws.cell(last_row + 3, 33).value = '*빨간음영 : 주력모델'  # 글자 데이터 입력
ws.cell(last_row + 3, 33).font = Font(color='FF0000', bold=True)  # 색상, 볼드체
ws.cell(last_row + 3, 33).alignment = align = Alignment(horizontal='right')  # 우측 정렬 




# 2번 시트 선택
ws = wb['2. IT 주력모델 시세현황']


# B열의 셀들 가져오기
cells = ws['B']


# 마지막행 B열 값 확인
last_row = ws.max_row

# B열 (넘버링) 위의 행 번호 + 1 입력 
last_value = ws.cell(row=last_row, column=2).value  # 마지막 값 불러오기
ws.cell(row=last_row+1, column=2, value=last_value + 1)

# M열 (넘버링) 위의 행 번호 + 1 입력
last_value = ws.cell(row=last_row, column=13).value
ws.cell(row=last_row+1, column=13, value=last_value + 1)

for i in range(len(d)):
    ws.cell(last_row + 1, i+3).value = d[i]
    
for i in range(len(e)):
    ws.cell(last_row + 1, i+14).value = e[i]
    

for k in [4,6,8,15,17]:
    # k열의 마지막행과 마지막행의 윗행 값 획득
    last_row_val = ws.cell(row=last_row+1, column=k).value
    prev_row_val = ws.cell(row=last_row, column=k).value

    # 마지막 행의 k+1 열에 계산값(변동액) 입력
    ws.cell(row=last_row+1, column= k+1).value = last_row_val - prev_row_val


# 변경된 내용을 저장
current_dir = os.getcwd()
file_name = f'{cr_today}_IT 시세 취합.xlsx'
file_path = os.path.join(current_dir, file_name)



wb.save(file_path)

