# 모듈(엑셀 후처리)
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from crawling import *

def load_file(srh_krd):
    dir = f'C:/flyordig/bidding_crawling/{crawling_date}-{srh_krd}검색.xlsx'
    wb = load_workbook(dir) # 해당 경로의 엑셀 파일 불러오기
    ws = wb.active # 불러온 엑셀 파일의 시트 활성화